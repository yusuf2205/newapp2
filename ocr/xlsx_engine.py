"""Извлечение данных из .xlsx — тот же договор/счёт, что и .docx, но в виде таблицы
Excel (встречается на практике параллельно с .docx-версией того же документа).

В отличие от .docx, где таблица-бланк обычно единственная и очевидная, лист Excel —
это одна большая сетка без явных границ таблиц. Поэтому ищем заголовок (строку с
≥2 узнаваемыми колонками) по всем строкам листа, а данные под ним берём до первой
полностью пустой строки.
"""

from __future__ import annotations

import asyncio
import io

from .base import OcrError, parse_plain_text
from .table_utils import goods_from_items_grid, table_fields_from_grid

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _load_workbook(content: bytes):
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise OcrError("Установите openpyxl") from exc

    try:
        return openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise OcrError(f"Не удалось прочитать .xlsx: {exc}") from exc


def _sheet_rows(worksheet) -> list[list[str]]:
    rows = []
    for row in worksheet.iter_rows(values_only=True):
        rows.append(["" if value is None else str(value) for value in row])
    return rows


def _is_blank_row(row: list[str]) -> bool:
    return not any(cell.strip() for cell in row)


def _flatten_text(rows: list[list[str]]) -> str:
    parts = []
    for row in rows:
        cells = [cell.strip() for cell in row if cell.strip()]
        if cells:
            parts.append(" | ".join(cells))
    return "\n".join(parts)


def _extract_from_sheet(rows: list[list[str]]) -> dict:
    result: dict = {}
    for index, row in enumerate(rows):
        if _is_blank_row(row):
            continue

        data_rows = []
        for candidate in rows[index + 1:]:
            if _is_blank_row(candidate):
                break
            data_rows.append(candidate)
        if not data_rows:
            continue

        grid = [row] + data_rows

        if "goods" not in result:
            goods = goods_from_items_grid(grid)
            if goods:
                result["goods"] = goods

        for key, value in table_fields_from_grid(grid).items():
            result.setdefault(key, value)

    return result


async def recognize_xlsx(content: bytes) -> dict:
    workbook = await asyncio.to_thread(_load_workbook, content)

    text_parts: list[str] = []
    table_result: dict = {}
    for worksheet in workbook.worksheets:
        rows = await asyncio.to_thread(_sheet_rows, worksheet)
        text_parts.append(_flatten_text(rows))
        for key, value in _extract_from_sheet(rows).items():
            table_result.setdefault(key, value)

    text = "\n".join(part for part in text_parts if part)
    if not text.strip():
        raise OcrError("В документе не найдено текста")

    result = parse_plain_text(text)
    result.update(table_result)
    return result
